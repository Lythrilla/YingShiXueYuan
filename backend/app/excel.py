import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import models

STATUS_LABELS = {"booked": "已预约", "verified": "已核销", "cancelled": "已取消"}

HEADERS = [
    "预约编号",
    "资源名称",
    "类型",
    "预约日期",
    "时间段",
    "预约人",
    "电话",
    "专业",
    "录音人数",
    "指导教师",
    "数量(套)",
    "状态",
    "提交时间",
    "核销时间",
    "录音事项说明",
]


def bookings_to_xlsx(bookings: list[models.Booking]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "预约报表"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, b in enumerate(bookings, start=2):
        slot_text = f"{b.slot.name} {b.slot.start_time}-{b.slot.end_time}" if b.slot else ""
        kind_text = "实验室" if b.resource and b.resource.kind == "lab" else "设备"
        values = [
            b.id,
            b.resource.name if b.resource else "",
            kind_text,
            b.date,
            slot_text,
            b.applicant_name,
            b.phone,
            b.major,
            b.num_people,
            b.instructor,
            b.quantity,
            STATUS_LABELS.get(b.status, b.status),
            b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
            b.verified_at.strftime("%Y-%m-%d %H:%M") if b.verified_at else "",
            b.description,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)

    widths = [10, 22, 8, 12, 18, 12, 16, 14, 10, 12, 10, 10, 18, 18, 40]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
